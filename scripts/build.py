"""
Build the DICE workbook.

"""
import os
import configparser
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

from extract import extract_data

CONFIG = configparser.ConfigParser()
CONFIG.read(os.path.join(os.path.dirname(__file__), 'script_config.ini'))
BASE_PATH = CONFIG['file_locations']['base_path']

DATA_RAW = os.path.join(BASE_PATH, 'raw')
DATA_INTERMEDIATE = os.path.join(BASE_PATH, 'intermediate')


def generate_workbook():
    """
    Generate the workbook and associated worksheets.

    """
    #import the Workbook class
    wb = Workbook()

    index = wb.active
    index = add_index(index)

    readme = wb.create_sheet("Read_Me", (2-1))
    readme = add_readme(readme)

    countries = wb.create_sheet("Countries", (3-1))
    countries = add_country_selection(countries)

    settings = wb.create_sheet("Settings", (4-1))
    settings = add_settings(settings)

    estimates = wb.create_sheet("Estimates", (5-1))
    estimates = add_estimates(estimates)

    population = wb.create_sheet("Pop", (6-1))
    population, lnth = add_country_data(population, 'population')

    cols = ['A','B','C','D','E','F','G','H','I','J','K','L']
    cum_pop = wb.create_sheet("Pop_C", (7-1))
    cum_pop = add_cum_pop(cum_pop, cols)

    area = wb.create_sheet("Area", (8-1))
    area, lnth = add_country_data(area, 'area_km2')

    pop_density = wb.create_sheet("P_Density_km2", (9-1))
    pop_density, lnth = add_country_data(pop_density, 'population_km2')

    pop_growth = wb.create_sheet("P_Growth", (10-1))
    pop_growth, lnth2 = add_pop_growth(pop_growth)

    users = wb.create_sheet("Users_km2", (11-1))
    users = add_users(users, cols, lnth)
    users.sheet_properties.tabColor = "66ff99"

    data_demand = wb.create_sheet("Data_km2", (12-1))
    data_demand = add_data_demand(data_demand, cols, lnth)
    data_demand.sheet_properties.tabColor = "66ff99"

    lookups = wb.create_sheet("Lookups", (13-1))
    lookups = add_lookups_sheet(lookups)
    lookups.sheet_properties.tabColor = "0000ff"

    coverage = wb.create_sheet("Coverage", (14-1))
    coverage = add_coverage_sheet(coverage, cols)
    coverage.sheet_properties.tabColor = "0000ff"

    towers = wb.create_sheet("Towers", (15-1))
    towers = add_towers_sheet(towers, cols, lnth)
    towers.sheet_properties.tabColor = "0000ff"

    towers_4G = wb.create_sheet("Towers_4G", (16-1))
    towers_4G = add_towers_4G_sheet(towers_4G, cols, lnth)
    towers_4G.sheet_properties.tabColor = "0000ff"

    towers_4G_km2 = wb.create_sheet("Towers_4G_km2", (17-1))
    towers_4G_km2 = add_towers_4G_km2_sheet(towers_4G_km2, cols, lnth)
    towers_4G_km2.sheet_properties.tabColor = "0000ff"

    towers_mno = wb.create_sheet("Towers_MNO", (18-1))
    towers_mno = add_towers_mno_sheet(towers_mno, cols, lnth)
    towers_mno.sheet_properties.tabColor = "0000ff"

    towers_4G_mno = wb.create_sheet("Towers_4G_MNO", (19-1))
    towers_4G_mno = add_towers_4G_mno_sheet(towers_4G_mno, cols, lnth)
    towers_4G_mno.sheet_properties.tabColor = "0000ff"

    towers_non_4G_mno = wb.create_sheet("Towers_non_4G_MNO", (20-1))
    towers_non_4G_mno = add_towers_non_4G_mno_sheet(towers_non_4G_mno, cols, lnth)
    towers_non_4G_mno.sheet_properties.tabColor = "0000ff"

    capacity = wb.create_sheet("Capacity_km2_MNO", (21-1))
    capacity = add_capacity_sheet(capacity, cols, lnth)
    capacity.sheet_properties.tabColor = "0000ff"

    sites = wb.create_sheet("Total_Sites_MNO", (22-1))
    sites = add_sites_sheet(sites, cols, lnth)
    sites.sheet_properties.tabColor = "0000ff"

    site_users = wb.create_sheet("Site_Users", (23-1))
    site_users = add_site_users_sheet(site_users, cols, lnth)
    site_users.sheet_properties.tabColor = "0000ff"

    sites_km2 = wb.create_sheet("Total_Sites_km2", (24-1))
    sites_km2 = add_sites_km2_sheet(sites_km2, cols, lnth)
    sites_km2.sheet_properties.tabColor = "0000ff"

    new = wb.create_sheet("New_4G_Sites", (25-1))
    new = add_new_sites_sheet(new, cols, lnth)
    new.sheet_properties.tabColor = "0000ff"

    upgrades = wb.create_sheet("Upgrades", (26-1))
    upgrades = add_upgrades(upgrades, cols, lnth)
    upgrades.sheet_properties.tabColor = "0000ff"

    new_builds = wb.create_sheet("New_Builds", (27-1))
    new_builds = add_new_builds(new_builds, cols, lnth)
    new_builds.sheet_properties.tabColor = "0000ff"

    site_costs = wb.create_sheet("RAN_Capex", (28-1))
    site_costs = add_site_costs(site_costs, cols, lnth)

    bh_costs = wb.create_sheet("BH_Capex", (29-1))
    bh_costs = add_bh_costs(bh_costs, cols, lnth)

    tower_costs = wb.create_sheet("Tower_Capex", (30-1))
    tower_costs = add_tower_costs(tower_costs, cols, lnth)

    labor_costs = wb.create_sheet("Labor_Capex", (31-1))
    labor_costs = add_labor_costs(labor_costs, cols, lnth)

    power_cost = wb.create_sheet("Power_Capex", (32-1))
    power_cost = add_power_costs(power_cost, cols, lnth)

    site_opex = wb.create_sheet("RAN_Opex", (33-1))
    site_opex = add_site_opex(site_opex, cols, lnth)

    bh_opex = wb.create_sheet("BH_Opex", (34-1))
    bh_opex = add_bh_opex(bh_opex, cols, lnth)

    tower_opex = wb.create_sheet("Tower_Opex", (35-1))
    tower_opex = add_tower_opex(tower_opex, cols, lnth)

    power_opex = wb.create_sheet("Power_Opex", (36-1))
    power_opex = add_power_opex(power_opex, cols, lnth)

    mno_cost = wb.create_sheet("MNO_Costs", (37-1))
    mno_cost = add_mno_costs(mno_cost, cols, lnth)

    total_cost = wb.create_sheet("Total_Costs", (38-1))
    total_cost = add_total_costs(total_cost, cols, lnth)

    gdp = wb.create_sheet("GDP", (39-1))
    gdp = add_gdp_sheet(gdp)

    options = wb.create_sheet("Options", (40-1))
    options = add_options(options)

    # ################
    # ###create graphs

    # estimates = relocate_context_data(estimates)

    context_data = wb.create_sheet("Context_Data", (41-1))
    add_context_data(context_data, population, area, pop_density,
        pop_growth, users, data_demand)

    estimates = relocate_estimates_data(estimates)

    context = wb.create_sheet("Country_Context", (42-1))
    add_country_context(context, context_data)

    context = wb.create_sheet("Country_Costs", (43-1))
    add_country_costs(context, estimates)

    context = wb.create_sheet("Income_Group_Costs", (44-1))
    add_income_group_costs(context, estimates)

    context = wb.create_sheet("Regional_Costs", (45-1))
    add_regional_group_costs(context, estimates)

    estimates.sheet_state = 'hidden'
    population.sheet_state  = 'hidden'
    area.sheet_state = 'hidden'
    pop_density.sheet_state = 'hidden'
    pop_growth.sheet_state = 'hidden'
    cum_pop.sheet_state = 'hidden'
    users.sheet_state = 'hidden'
    data_demand.sheet_state = 'hidden'
    lookups.sheet_state = 'hidden'
    coverage.sheet_state = 'hidden'
    towers.sheet_state = 'hidden'
    towers_4G.sheet_state = 'hidden'
    towers_4G_km2.sheet_state = 'hidden'
    towers_mno.sheet_state = 'hidden'
    towers_4G_mno.sheet_state = 'hidden'
    towers_non_4G_mno.sheet_state = 'hidden'
    site_users.sheet_state = 'hidden'
    capacity.sheet_state = 'hidden'
    sites.sheet_state = 'hidden'
    sites_km2.sheet_state = 'hidden'
    new.sheet_state = 'hidden'
    upgrades.sheet_state = 'hidden'
    new_builds.sheet_state = 'hidden'
    site_costs.sheet_state = 'hidden'
    bh_costs.sheet_state = 'hidden'
    tower_costs.sheet_state = 'hidden'
    labor_costs.sheet_state = 'hidden'
    power_cost.sheet_state = 'hidden'
    site_opex.sheet_state = 'hidden'
    bh_opex.sheet_state = 'hidden'
    tower_opex.sheet_state = 'hidden'
    power_opex.sheet_state = 'hidden'
    mno_cost.sheet_state = 'hidden'
    total_cost.sheet_state = 'hidden'
    gdp.sheet_state = 'hidden'
    options.sheet_state = 'hidden'
    context_data.sheet_state = 'hidden'

    wb.save('Oughton et al. (2022) DICE.xlsx')

    return print("Generated workbook")


def add_index(ws):
    """
    Add the welcome index page.

    """
    ws.title = "Index"
    ws.sheet_properties.tabColor = "004C97"

    ws = set_font(ws, 'A1:AZ1000', 'Segoe UI')

    #Set height of row one
    ws.row_dimensions[1].height = 34.5
    ws.row_dimensions[2].height = 14.25

    ##Color white
    ws.sheet_view.showGridLines = False
    # set_border(ws, 'A1:AZ1000', "thin", "00FFFFFF")

    #Set blue and red border strips
    set_cell_color(ws, 'A1:AZ1', "004C97")
    set_cell_color(ws, 'A2:AZ2', "C00000")

    ws['B4'] = "IMF FAD Digital Infrastructure Costing Estimator (DICE)"
    ws['B4'].font = Font(size=20)
    ws['B5'] = "Developed by George Mason University and the Fiscal Affairs Department (FAD), International Monetary Fund (IMF)"

    path = os.path.join(BASE_PATH, '..', 'images', 'imf_logo.png')
    img = Image(path)
    img.height = 150 # insert image height in pixels as float or int (e.g. 305.5)
    img.width = 150# insert image width in pixels as float or int (e.g. 405.8)
    img.anchor = 'B7'
    ws.add_image(img)

    ws['B15'] = "Contact: Edward Oughton (eoughton@gmu.edu); David Amaglobeli (damaglobeli@imf.org); Mariano Moszoro (mmoszoro@imf.org)."
    ws['B16'] = '=HYPERLINK("{}")'.format('https://github.com/edwardoughton/dice')
    hyperlink = Font(underline='single', color='0563C1')
    ws['B16'].font = hyperlink

    ws['B18'] = "Capabilities/Uses:"
    ws['B18'].font = Font(bold=True)
    ws['B19'] = """    - Supporting high-level decisions pertaining to universal broadband investment by capturing the main cost drivers which affect the deployment of broadband infrastructure."""
    ws['B20'] = "    - Breaking down investment by specific country income groups and regions."

    ws['B22'] = "Caveats to using DICE:"
    ws['B22'].font = Font(bold=True)
    ws['B23'] = """    - DICE is not a replacement for detailed country-specific modeling. The aim is to provide comparative understanding across all countries globally."""
    ws['B24'] = """    - DICE is not an exact measurement tool. Accuracy and precision are commensurate with being able to make global cross-country comparisons."""
    ws['B25'] = """    - DICE is not a stand-alone policy option tool. If deep subject matter exertise are required, country teams should reach out to the DICE development team at GMU and the IMF (eoughton@gmu.edu; damaglobeli@imf.org; mmoszoro@imf.org)."""

    ws['B27'] = "Main Datasets:"
    ws['B27'].font = Font(bold=True)
    ws['B28'] = "WorldPop 2020 Unconstrained 1km Global Mosaic, UN Population Forecasts 2020-2030, IMF GDP Projections 2020-2030"

    ws['B30'] = "Reference:"
    ws['B30'].font = Font(bold=True)
    ws['B31'] = "    - Working Paper Citation TBC."
    return ws


def add_readme(ws):
    """

    """
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "004C97"
    set_cell_color(ws, 'A1:Q6', "004C97")
    ws = set_font(ws, 'A1:AZ1000', 'Segoe UI')

    path = os.path.join(BASE_PATH, '..', 'images', 'imf_logo.png')
    img = Image(path)
    img.height = 133 # insert image height in pixels as float or int (e.g. 305.5)
    img.width = 133 # insert image width in pixels as float or int (e.g. 405.8)
    img.anchor = 'A1'
    ws.add_image(img)

    ws.merge_cells('A3:Q4')
    ws['A3'] = "Read Me"
    ws['A3'].font = Font(size=20, color='FFFFFF')
    ws['A3'].alignment = Alignment(horizontal='center')

    ws['B9'] = 'Overview'
    ws['B10'] = '- This tool provides investment analytics for achieving universal mobile broadband (SDG Target 9c).'
    ws['B11'] = '- DICE complements the IMF’s other SDG infrastructure costing tools, covering roads, electricity, and water and sanitation.'

    ws['B13'] = 'Approach'
    ws['B14'] = '- Comparative country-specific investment needs can be estimated to achieve universal broadband connectivity using 4G.'
    ws['B15'] = '- The focus is on wireless cellular connectivity as this is one of the cheapest ways to provide wide-area broadband services affordably.'
    ws['B16'] = '- The level of universal broadband can be specified by the DICE user, based on either a minimum speed or monthly data consumption.'
    ws['B17'] = '- DICE estimates account for demographic forecast trends, population density, and economic characteristics.'

    ws['B19'] = 'The DICE Method'
    path = os.path.join(BASE_PATH, '..', 'images', 'method.png')
    img = Image(path)
    img.anchor = 'B20'
    ws.add_image(img)

    ws['B9'].font = Font(bold=True, name='Segoe UI')
    ws['B13'].font = Font(bold=True, name='Segoe UI')
    ws['B19'].font = Font(bold=True, name='Segoe UI')

    return ws


def add_settings(ws):
    """

    """
    ##Color white
    ws.sheet_view.showGridLines = False
    #Set blue and red border strips
    set_cell_color(ws, 'A1:AZ1', "004C97")
    set_cell_color(ws, 'A2:AZ2', "C00000")

    ##Allocate title
    ws.title = "Settings"
    ws.sheet_properties.tabColor = "004C97"

    # ##Set column width
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40

    ## Add parameters box
    ws.merge_cells('B6:C6')
    ws['B6'] = "Time Parameters"

    ws['B7'] = "Parameter"
    ws['C7'] = "Option"

    ws['B8'] = "Start Year"
    ws['C8'] = 2022

    ws['B9'] = "End Year"
    ws['C9'] = 2030

    ws['B10'] = "Total Years"
    ws['C10'] = "=C9-C8"

    ws['B11'] = "Depreciation"
    ws['C11'] = "=0.08"

    ws.merge_cells('B13:C13')
    ws['B13'] = "Demand Parameters"

    ws['B14'] = "Parameter"
    ws['C14'] = "Option"

    ws['B15'] = "Smartphone adoption (%)"
    ws['C15'] = 95

    ws['B16'] = "Market Share (%)"
    ws['C16'] = 25

    ws['B17'] = "Active Users (%)"
    ws['C17'] = 5

    ws['B18'] = "Minimum capacity per user (Mbps)"
    ws['C18'] = 1

    ws['B19'] = "Data demand per month (GB)"
    ws['C19'] = 90

    ws.merge_cells('B21:C21')
    ws['B21'] = "Supply Parameters"

    ws['B22'] = "Parameter"
    ws['C22'] = "Option"

    ws['B23'] = "Infrastructure Strategy"
    data_val = DataValidation(type="list", formula1='=Options!D2:D6')
    ws.add_data_validation(data_val)
    data_val.add(ws["C23"])
    ws['C23'] = "4G"

    ws['B24'] = "Existing Spectrum Availability"
    data_val = DataValidation(type="list", formula1='=Options!E2:E4')
    ws.add_data_validation(data_val)
    data_val.add(ws["C24"])
    ws['C24'] = "Baseline"

    ws['B25'] = "Future Spectrum Availability"
    data_val = DataValidation(type="list", formula1='=Options!E2:E4')
    ws.add_data_validation(data_val)
    data_val.add(ws["C25"])
    ws['C25'] = "Baseline"

    ws['B26'] = "Minimum Pop.Density to Serve (km^2)"
    ws['C26'] = 0

    format_numbers(ws, ['C'], (11,11), 'Percent', 1)

    #Center text
    ws = center_text(ws, 'A2:AZ1000')
    # set_border(ws, 'A1:AZ1000', "thin", "00FFFFFF")
    ws = set_font(ws, 'A1:AZ1000', 'Segoe UI')
    ws = set_bold(ws, 'B6', 'Segoe UI')
    ws = set_bold(ws, 'B7', 'Segoe UI')
    ws = set_bold(ws, 'C7', 'Segoe UI')
    ws = set_bold(ws, 'B13', 'Segoe UI')
    ws = set_bold(ws, 'B14', 'Segoe UI')
    ws = set_bold(ws, 'C14', 'Segoe UI')
    ws = set_bold(ws, 'B21', 'Segoe UI')
    ws = set_bold(ws, 'B22', 'Segoe UI')
    ws = set_bold(ws, 'C22', 'Segoe UI')

    set_border(ws, 'B6:C11', "thin", "000000")
    set_border(ws, 'B13:C19', "thin", "000000")
    set_border(ws, 'B21:C26', "thin", "000000")

    yellow_fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
    ws['C8'].fill = yellow_fill
    ws['C9'].fill = yellow_fill
    ws['C10'].fill = yellow_fill
    ws['C11'].fill = yellow_fill

    ws['C15'].fill = yellow_fill
    ws['C16'].fill = yellow_fill
    ws['C17'].fill = yellow_fill
    ws['C18'].fill = yellow_fill
    ws['C19'].fill = yellow_fill

    ws['C23'].fill = yellow_fill
    ws['C24'].fill = yellow_fill
    ws['C25'].fill = yellow_fill
    ws['C26'].fill = yellow_fill

    ws.merge_cells('B28:C28')
    ws['B28'] = 'Yellow cells are capable of user-defined specification.'
    ws = set_bold(ws, 'B28', 'Segoe UI')

    return ws


def add_country_selection(ws):
    """
    Add country selection.

    """
    ##Color white
    ws.sheet_view.showGridLines = False
    #Set blue and red border strips
    set_cell_color(ws, 'A1:AZ1', "004C97")
    set_cell_color(ws, 'A2:AZ2', "C00000")

    ##Allocate title
    ws.title = "Countries"
    ws.sheet_properties.tabColor = "004C97"

    ## Set widths
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40

    yellow_fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
    ws['B8'].fill = yellow_fill
    ws['B9'].fill = yellow_fill
    ws['B10'].fill = yellow_fill
    ws['B11'].fill = yellow_fill

    ## Add parameters box
    format_numbers(ws, ['C'], (11,11), 'Percent', 0)
    set_border(ws, 'B6:C11', "thin", "000000")
    ws.merge_cells('B6:C6')
    ws['B6'] = "Select Countries For Comparison"
    ws['B7'] = "Country"
    ws['C7'] = "ISO3"

    ### Country 1
    data_val = DataValidation(type="list", formula1='=Options!A2:A251')
    ws.add_data_validation(data_val)
    data_val.add(ws["B8"])
    ws['B8'] = "Afghanistan"
    ws['C8'] = "=IFERROR(INDEX(Options!B2:B1611,MATCH(B8, Options!A2:A1611)), \"\")"

    # ### Country 2
    data_val = DataValidation(type="list", formula1='=Options!A2:A251')
    ws.add_data_validation(data_val)
    data_val.add(ws["B9"])
    ws['B9'] = "Bhutan"
    ws['C9'] = "=IFERROR(INDEX(Options!B2:B1611,MATCH(B9, Options!A2:A1611)), \"\")"

    ### Country 3
    data_val = DataValidation(type="list", formula1='=Options!A2:A251')
    ws.add_data_validation(data_val)
    data_val.add(ws["B10"])
    ws['B10'] = "Bangladesh"
    ws['C10'] = "=IFERROR(INDEX(Options!B2:B1611,MATCH(B10, Options!A2:A1611)), \"\")"

    ### Country 4
    data_val = DataValidation(type="list", formula1='=Options!A2:A251')
    ws.add_data_validation(data_val)
    data_val.add(ws["B11"])
    ws['B11'] = "India"
    ws['C11'] = "=IFERROR(INDEX(Options!B2:B1611,MATCH(B11, Options!A2:A1611)), \"\")"

    ws.merge_cells('B13:C13')
    ws.merge_cells('B14:C14')
    ws['B13'] = 'Select the yellow cells to access the drop-down list of countries.'
    ws['B14'] = 'Up to four countries can then be selected for comparison.'

    #Center text
    ws = center_text(ws, 'A2:AZ1000')
    # set_border(ws, 'A1:AZ1000', "thin", "00FFFFFF")
    ws = set_font(ws, 'A1:AZ1000', 'Segoe UI')
    ws = set_bold(ws, 'B6', 'Segoe UI')
    ws = set_bold(ws, 'B7', 'Segoe UI')
    ws = set_bold(ws, 'C7', 'Segoe UI')
    ws = set_bold(ws, 'B13', 'Segoe UI')
    ws = set_bold(ws, 'B14', 'Segoe UI')

    return ws


def add_estimates(ws):
    """

    """
    ##Color white
    ws.sheet_view.showGridLines = False
    #Set blue and red border strips
    set_cell_color(ws, 'A1:AZ1', "004C97")
    set_cell_color(ws, 'A2:AZ2', "C00000")

    ##Allocate title
    # ws.title = "Settings"
    ws.sheet_properties.tabColor = "004C97"

    #Cross Country Comparisons
    ws.merge_cells('B6:I6')
    ws['B6'] = "Cross-Country Comparisons"
    ws['B7'] = "Country"
    ws['C7'] = "ISO3"
    ws['D7'] = "Total Cost ($Bn)"
    ws['E7'] = "Mean Annual 10-Year GDP ($Bn)"
    ws['F7'] = "GDP Growth Rate (%)(2022-2030)"
    ws['G7'] = "Initial Investment ($Bn)"
    ws['H7'] = "2022 GDP ($Bn)"
    ws['I7'] = "Annual GDP Share (%)"

    # ### Country 1
    ws['B8'] = "=IFERROR(Countries!B8, \"\")"
    ws['C8'] = "=IFERROR(Countries!C8, \"\")"
    ws['D8'] = "=IFERROR(INDEX(Total_Costs!M2:M1611,MATCH(C8, Total_Costs!A2:A1611))/1e9, \"\")"
    ws['E8'] = "=IFERROR(INDEX(GDP!L2:L1611,MATCH(C8, GDP!A2:A1611)), \"\")"
    ws['F8'] = "=IFERROR(INDEX(GDP!M2:M1611,MATCH(C8, GDP!A2:A1611)), )"
    ws['G8'] = "=IF(F8=0,D8/Settings!C10,(D8*(1-(1+F8)/(1-Settings!C11)))/((1-Settings!C11)^Settings!C10*(1-((1+F8)/(1-Settings!C11))^(Settings!C10+1))))"
    ws['H8'] = "=IFERROR(INDEX(GDP!C2:C1611,MATCH(C8, GDP!A2:A1611)), \"\")"
    ws['I8'] = "=IFERROR(G8/H8, \"\")"

    # ### Country 2
    ws['B9'] = "=IFERROR(Countries!B9, \"\")"
    ws['C9'] = "=IFERROR(Countries!C9, \"\")"
    ws['D9'] = "=IFERROR(INDEX(Total_Costs!M2:M1611,MATCH(C9, Total_Costs!A2:A1611))/1e9, \"\")"
    ws['E9'] = "=IFERROR(INDEX(GDP!L2:L1611,MATCH(C9, GDP!A2:A1611)), \"\")"
    ws['F9'] = "=IFERROR(INDEX(GDP!M2:M1611,MATCH(C9, GDP!A2:A1611)), )"
    ws['G9'] = "=IF(F8=0,D9/Settings!C10,(D9*(1-(1+F9)/(1-Settings!C11)))/((1-Settings!C11)^Settings!C10*(1-((1+F9)/(1-Settings!C11))^(Settings!C10+1))))"
    ws['H9'] = "=IFERROR(INDEX(GDP!C2:C1611,MATCH(C9, GDP!A2:A1611)), \"\")"
    ws['I9'] = "=IFERROR(G9/H9, \"\")"

    # ### Country 3
    ws['B10'] = "=IFERROR(Countries!B10, \"\")"
    ws['C10'] = "=IFERROR(Countries!C10, \"\")"
    ws['D10'] = "=IFERROR(INDEX(Total_Costs!M2:M1611,MATCH(C10, Total_Costs!A2:A1611))/1e9, \"\")"
    ws['E10'] = "=IFERROR(INDEX(GDP!L2:L1611,MATCH(C10, GDP!A2:A1611)), \"\")"
    ws['F10'] = "=IFERROR(INDEX(GDP!M2:M1611,MATCH(C10, GDP!A2:A1611)), )"
    ws['G10'] = "=IF(F8=0,D10/Settings!C10,(D10*(1-(1+F10)/(1-Settings!C11)))/((1-Settings!C11)^Settings!C10*(1-((1+F10)/(1-Settings!C11))^(Settings!C10+1))))"
    ws['H10'] = "=IFERROR(INDEX(GDP!C2:C1611,MATCH(C10, GDP!A2:A1611)), \"\")"
    ws['I10'] = "=IFERROR(G10/H10, \"\")"

    # ### Country 4
    ws['B11'] = "=IFERROR(Countries!B11, \"\")"
    ws['C11'] = "=IFERROR(Countries!C11, \"\")"
    ws['D11'] = "=IFERROR(INDEX(Total_Costs!M2:M1611,MATCH(C11, Total_Costs!A2:A1611))/1e9, \"\")"
    ws['E11'] = "=IFERROR(INDEX(GDP!L2:L1611,MATCH(C11, GDP!A2:A1611)), \"\")"
    ws['F11'] = "=IFERROR(INDEX(GDP!M2:M1611,MATCH(C11, GDP!A2:A1611)), )"
    ws['G11'] = "=IF(F8=0,D11/Settings!C10,(D11*(1-(1+F11)/(1-Settings!C11)))/((1-Settings!C11)^Settings!C10*(1-((1+F11)/(1-Settings!C11))^(Settings!C10+1))))"
    ws['H11'] = "=IFERROR(INDEX(GDP!C2:C1611,MATCH(C11, GDP!A2:A1611)), \"\")"
    ws['I11'] = "=IFERROR(G11/H11, \"\")"

    # ##Set column width
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 30

    ws = format_numbers(ws, ['D', 'E', 'G', 'H'], (8,11), 'Comma [0]', 1)
    ws = format_numbers(ws, ['D', 'E', 'G', 'H'], (17,19), 'Comma [0]', 1)
    ws = format_numbers(ws, ['D', 'E', 'G', 'H'], (24,30), 'Comma [0]', 1)
    ws = format_numbers(ws, ['F'], (8,11), 'Percent', 1)
    ws = format_numbers(ws, ['F'], (17,19), 'Percent', 1)
    ws = format_numbers(ws, ['F'], (24,30), 'Percent', 1)
    ws = format_numbers(ws, ['I'], (8,11), 'Percent', 2)
    ws = format_numbers(ws, ['I'], (17,19), 'Percent', 2)
    ws = format_numbers(ws, ['I'], (24,30), 'Percent', 2)

    # ### Costs by Income Group
    ws.merge_cells('B15:I15')
    ws['B15'] = "Cost by Income Group"
    ws.merge_cells('B16:C16')
    ws['B16'] = "Income Group"
    ws['D16'] = "Total Cost ($Bn)"
    ws['E16'] = "Mean Annual 10-Year GDP ($Bn)"
    ws['F16'] = "GDP Growth Rate (%)(2022-2030)"
    ws['G16'] = "Initial Investment ($Bn)"
    ws['H16'] = "2022 GDP ($Bn)"
    ws['I16'] = "Annual GDP Share (%)"

    ws.merge_cells('B17:C17')
    ws['B17'] = 'Advanced Economies'
    ws['D17'] = '=SUMIF(Total_Costs!$O$2:$O$250,B17,Total_Costs!$M$2:$M$250)/1e9'
    ws['E17'] = '=SUMIF(GDP!$N$2:$N$250,B17,GDP!$L$2:$L$250)'
    ws['F17'] = '=AVERAGEIF(GDP!$N$2:$N$250,B17,GDP!$M$2:$M$250)'
    ws['G17'] = "=IF(F17=0,D17/Settings!C10,(D17*(1-(1+F17)/(1-Settings!C11)))/((1-Settings!C11)^Settings!C10*(1-((1+F17)/(1-Settings!C11))^(Settings!C10+1))))"
    ws['H17'] = "=SUMIF(GDP!$N$2:$N$250,B17,GDP!$C$2:$C$250)"
    ws['I17'] = "=IFERROR(G17/H17, \"\")"

    ws.merge_cells('B18:C18')
    ws['B18'] = 'Emerging Market Economies'
    ws['D18'] = '=SUMIF(Total_Costs!$O$2:$O$250,B18,Total_Costs!$M$2:$M$250)/1e9'
    ws['E18'] = '=SUMIF(GDP!$N$2:$N$250,B18,GDP!$L$2:$L$250)'
    ws['F18'] = '=AVERAGEIF(GDP!$N$2:$N$250,B18,GDP!$M$2:$M$250)'
    ws['G18'] = "=IF(F18=0,D18/Settings!C10,(D18*(1-(1+F18)/(1-Settings!C11)))/((1-Settings!C11)^Settings!C10*(1-((1+F17)/(1-Settings!C11))^(Settings!C10+1))))"
    ws['H18'] = "=SUMIF(GDP!$N$2:$N$250,B18,GDP!$C$2:$C$250)"
    ws['I18'] = "=IFERROR(G18/H18, \"\")"

    ws.merge_cells('B19:C19')
    ws['B19'] = 'Low Income Developing Countries'
    ws['D19'] = '=SUMIF(Total_Costs!$O$2:$O$250,B19,Total_Costs!$M$2:$M$250)/1e9'
    ws['E19'] = '=SUMIF(GDP!$N$2:$N$250,B19,GDP!$L$2:$L$250)'
    ws['F19'] = '=AVERAGEIF(GDP!$N$2:$N$250,B19,GDP!$M$2:$M$250)'
    ws['G19'] = "=IF(F19=0,D19/Settings!C10,(D19*(1-(1+F19)/(1-Settings!C11)))/((1-Settings!C11)^Settings!C10*(1-((1+F17)/(1-Settings!C11))^(Settings!C10+1))))"
    ws['H19'] = "=SUMIF(GDP!$N$2:$N$250,B19,GDP!$C$2:$C$250)"
    ws['I19'] = "=IFERROR(G19/H19, \"\")"

    # ### Costs by Region
    ws.merge_cells('B22:I22')
    ws['B22'] = "Cost by Region"
    ws.merge_cells('B23:C23')
    ws['B23'] = "Region"
    ws['D23'] = "Total Cost ($Bn)"
    ws['E23'] = "Mean Annual 10-Year GDP ($Bn)"
    ws['F23'] = "GDP Growth Rate (%)(2022-2030)"
    ws['G23'] = "Initial Investment ($Bn)"
    ws['H23'] = "2022 GDP ($Bn)"
    ws['I23'] = "Annual GDP Share (%)"

    ws.merge_cells('B24:C24')
    ws['B24'] = 'Advanced Economies'
    ws['D24'] = '=SUMIF(Total_Costs!$P$2:$P$250,B24,Total_Costs!$M$2:$M$250)/1e9'
    ws['E24'] = '=SUMIF(GDP!$O$2:$O$250,B24,GDP!$L$2:$L$250)'
    ws['F24'] = '=AVERAGEIF(GDP!$O$2:$O$250,B24,GDP!$M$2:$M$250)'
    ws['G24'] = "=IF(F24=0,D24/Settings!C17,(D24*(1-(1+F24)/(1-Settings!C11)))/((1-Settings!C11)^Settings!C17*(1-((1+F24)/(1-Settings!C11))^(Settings!C10+1))))"
    ws['H24'] = "=SUMIF(GDP!$O$2:$O$250,B24,GDP!$C$2:$C$250)"
    ws['I24'] = "=IFERROR(G24/H24, "")"

    ws.merge_cells('B25:C25')
    ws['B25'] = "Caucasus and Central Asia"
    ws['D25'] = '=SUMIF(Total_Costs!$P$2:$P$250,B25,Total_Costs!$M$2:$M$250)/1e9'
    ws['E25'] = '=SUMIF(GDP!$O$2:$O$250,B25,GDP!$L$2:$L$250)'
    ws['F25'] = '=AVERAGEIF(GDP!$O$2:$O$250,B25,GDP!$M$2:$M$250)'
    ws['G25'] = "=IF(F25=0,D25/Settings!C17,(D25*(1-(1+F25)/(1-Settings!C11)))/((1-Settings!C11)^Settings!C17*(1-((1+F25)/(1-Settings!C11))^(Settings!C10+1))))"
    ws['H25'] = "=SUMIF(GDP!$O$2:$O$250,B25,GDP!$C$2:$C$250)"
    ws['I25'] = "=IFERROR(G25/H25, "")"

    ws.merge_cells('B26:C26')
    ws['B26'] = "Emerging and Developing Asia"
    ws['D26'] = '=SUMIF(Total_Costs!$P$2:$P$250,B26,Total_Costs!$M$2:$M$250)/1e9'
    ws['E26'] = '=SUMIF(GDP!$O$2:$O$250,B26,GDP!$L$2:$L$250)'
    ws['F26'] = '=AVERAGEIF(GDP!$O$2:$O$250,B26,GDP!$M$2:$M$250)'
    ws['G26'] = "=IF(F26=0,D26/Settings!C17,(D26*(1-(1+F26)/(1-Settings!C11)))/((1-Settings!C11)^Settings!C17*(1-((1+F26)/(1-Settings!C11))^(Settings!C10+1))))"
    ws['H26'] = "=SUMIF(GDP!$O$2:$O$250,B26,GDP!$C$2:$C$250)"
    ws['I26'] = "=IFERROR(G26/H26, "")"

    ws.merge_cells('B27:C27')
    ws['B27'] = "Emerging and Developing Europe"
    ws['D27'] = '=SUMIF(Total_Costs!$P$2:$P$250,B27,Total_Costs!$M$2:$M$250)/1e9'
    ws['E27'] = '=SUMIF(GDP!$O$2:$O$250,B27,GDP!$L$2:$L$250)'
    ws['F27'] = '=AVERAGEIF(GDP!$O$2:$O$250,B27,GDP!$M$2:$M$250)'
    ws['G27'] = "=IF(F27=0,D27/Settings!C17,(D27*(1-(1+F27)/(1-Settings!C11)))/((1-Settings!C11)^Settings!C17*(1-((1+F27)/(1-Settings!C11))^(Settings!C10+1))))"
    ws['H27'] = "=SUMIF(GDP!$O$2:$O$250,B27,GDP!$C$2:$C$250)"
    ws['I27'] = "=IFERROR(G27/H27, "")"

    ws.merge_cells('B28:C28')
    ws['B28'] = "Latin America and the Caribbean"
    ws['D28'] = '=SUMIF(Total_Costs!$P$2:$P$250,B28,Total_Costs!$M$2:$M$250)/1e9'
    ws['E28'] = '=SUMIF(GDP!$O$2:$O$250,B28,GDP!$L$2:$L$250)'
    ws['F28'] = '=AVERAGEIF(GDP!$O$2:$O$250,B28,GDP!$M$2:$M$250)'
    ws['G28'] = "=IF(F28=0,D28/Settings!C17,(D28*(1-(1+F28)/(1-Settings!C11)))/((1-Settings!C11)^Settings!C17*(1-((1+F28)/(1-Settings!C11))^(Settings!C10+1))))"
    ws['H28'] = "=SUMIF(GDP!$O$2:$O$250,B28,GDP!$C$2:$C$250)"
    ws['I28'] = "=IFERROR(G28/H28, "")"

    ws.merge_cells('B29:C29')
    ws['B29'] = "Middle East, North Africa, Afghanistan, and Pakistan"
    ws['D29'] = '=SUMIF(Total_Costs!$P$2:$P$250,B29,Total_Costs!$M$2:$M$250)/1e9'
    ws['E29'] = '=SUMIF(GDP!$O$2:$O$250,B29,GDP!$L$2:$L$250)'
    ws['F29'] = '=AVERAGEIF(GDP!$O$2:$O$250,B29,GDP!$M$2:$M$250)'
    ws['G29'] = "=IF(F29=0,D29/Settings!C17,(D29*(1-(1+F29)/(1-Settings!C11)))/((1-Settings!C11)^Settings!C17*(1-((1+F29)/(1-Settings!C11))^(Settings!C10+1))))"
    ws['H29'] = "=SUMIF(GDP!$O$2:$O$250,B29,GDP!$C$2:$C$250)"
    ws['I29'] = "=IFERROR(G29/H29, "")"

    ws.merge_cells('B30:C30')
    ws['B30'] = "Sub-Sahara Africa"
    ws['D30'] = '=SUMIF(Total_Costs!$P$2:$P$250,B30,Total_Costs!$M$2:$M$250)/1e9'
    ws['E30'] = '=SUMIF(GDP!$O$2:$O$250,B30,GDP!$L$2:$L$250)'
    ws['F30'] = '=AVERAGEIF(GDP!$O$2:$O$250,B30,GDP!$M$2:$M$250)'
    ws['G30'] = "=IF(F30=0,D30/Settings!C17,(D30*(1-(1+F30)/(1-Settings!C11)))/((1-Settings!C11)^Settings!C17*(1-((1+F30)/(1-Settings!C11))^(Settings!C10+1))))"
    ws['H30'] = "=SUMIF(GDP!$O$2:$O$250,B30,GDP!$C$2:$C$250)"
    ws['I30'] = "=IFERROR(G30/H30, "")"

    set_border(ws, 'B6:I11', "thin", "000000")
    set_border(ws, 'B15:I19', "thin", "000000")
    set_border(ws, 'B22:I30', "thin", "000000")

    return ws


def relocate_estimates_data(ws):
    """

    """
    ### Country group
    ws = relocate(ws, 'B', 7, 11, 'B', 34)
    ws = relocate(ws, 'D', 7, 11, 'C', 34)
    ws = format_numbers(ws, ['C'], (34,40), 'Comma [0]', 3)

    ws = relocate(ws, 'B', 7, 11, 'E', 34)
    ws = relocate(ws, 'E', 7, 11, 'F', 34)
    ws = format_numbers(ws, ['F'], (34,40), 'Comma [0]', 0)

    ws = relocate(ws, 'B', 7, 11, 'H', 34)
    ws = relocate(ws, 'G', 7, 11, 'I', 34)
    ws = format_numbers(ws, ['I'], (34,40), 'Comma [0]', 3)

    ws = relocate(ws, 'B', 7, 11, 'B', 42)
    ws = relocate(ws, 'I', 7, 11, 'C', 42)
    ws = format_numbers(ws, ['C'], (42,48), 'Percent', 3)

    ### Income group
    ws = relocate(ws, 'B', 16, 19, 'B', 50)
    ws = relocate(ws, 'D', 16, 19, 'C', 50)
    ws = format_numbers(ws, ['C'], (51,53), 'Comma [0]', 3)

    ws = relocate(ws, 'B', 16, 19, 'E', 50)
    ws = relocate(ws, 'E', 16, 19, 'F', 50)
    ws = format_numbers(ws, ['F'], (51,53), 'Comma [0]', 0)

    ws = relocate(ws, 'B', 16, 19, 'H', 50)
    ws = relocate(ws, 'G', 16, 19, 'I', 50)
    ws = format_numbers(ws, ['I'], (51,53), 'Comma [0]', 3)

    ws = relocate(ws, 'B', 16, 19, 'B', 57)
    ws = relocate(ws, 'I', 16, 19, 'C', 57)
    ws = format_numbers(ws, ['C'], (58,60), 'Percent', 3)

    ### Region group
    ws = relocate(ws, 'B', 23, 30, 'B', 64)
    ws = relocate(ws, 'D', 23, 30, 'C', 64)
    ws = format_numbers(ws, ['C'], (65,72), 'Comma [0]', 3)

    ws = relocate(ws, 'B', 23, 30, 'E', 64)
    ws = relocate(ws, 'E', 23, 30, 'F', 64)
    ws = format_numbers(ws, ['F'], (65,72), 'Comma [0]', 0)

    ws = relocate(ws, 'B', 23, 30, 'H', 64)
    ws = relocate(ws, 'G', 23, 30, 'I', 64)
    ws = format_numbers(ws, ['I'], (65,72), 'Comma [0]', 3)

    ws = relocate(ws, 'B', 23, 30, 'B', 75)
    ws = relocate(ws, 'I', 23, 30, 'C', 75)
    ws = format_numbers(ws, ['C'], (76,83), 'Percent', 3)

    return ws


def relocate(ws, col, min_row, max_row, end_col, end_min_row):
    """
    ws, col, min_row, max_row, end_col, end_min_row
    worksheet
    starting_column
    starting_minimum_row
    starting_maximum_row
    ending_column
    ending_minimum_row

    """
    for i in range(min_row, max_row+1):
        starting_loc = '={}{}'.format(col, i)
        difference = end_min_row - min_row
        ending_loc = '{}{}'.format(end_col, i+difference)
        ws[ending_loc] = starting_loc #'=ROUND({},3-(1+INT(LOG10(ABS({})))))'.format(starting_loc, starting_loc)

    return ws


def format_numbers(ws, columns, set_range, format_type, number_format):

    lower, upper = set_range

    for column in columns:
        for i in range(lower, upper+1):
            cell = '{}{}'.format(column, i)

            ws[cell].style = format_type

            if format_type == 'Percent' and number_format == 1:
                ws[cell].number_format = '0.0%'
            elif format_type == 'Percent' and number_format == 2:
                ws[cell].number_format = '0.00%'
            elif format_type == 'Percent' and number_format == 3:
                ws[cell].number_format = '0.000%'
            elif format_type == 'Percent' and number_format == 4:
                ws[cell].number_format = '0.0000%'
            elif format_type == 'Percent' and number_format == 5:
                ws[cell].number_format = '0.00000%'
            elif number_format == 0:
                ws[cell].number_format = '0'
            elif number_format == 1:
                ws[cell].number_format = '0.0'
            elif number_format == 2:
                ws[cell].number_format = '0.00'
            elif number_format == 3:
                ws[cell].number_format = '0.000'
            elif number_format == 4:
                ws[cell].number_format = '0.0000'
            elif number_format == 5:
                ws[cell].number_format = '0.00000'

            ws[cell].font = Font(size=11)

    return ws


def set_border(ws, cell_range, style, color):

    thin = Side(border_style=style, color=color)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def set_cell_color(ws, cell_range, color):

    fill_gen = PatternFill(fill_type='solid',
                                 start_color=color,
                                 end_color=color)

    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill_gen


def set_font(ws, cell_range, font):
    """

    """
    for row in ws[cell_range]:
        for cell in row:
            cell.font = Font(name=font, size=11)

    return ws


def set_bold(ws, cell_id, font):
    """

    """
    # for row in ws[cell_range]:
    # for cell in ws[cell_range]:
    ws[cell_id].font = Font(name=font, size=11, bold=True)

    return ws


def center_text(ws, cell_range):
    """

    """
    for row in ws[cell_range]:
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    return ws


def add_options(ws):
    """

    """
    ws.sheet_properties.tabColor = "0099ff"

    path = os.path.join(BASE_PATH, 'global_information.csv')
    data = pd.read_csv(path, encoding = "ISO-8859-1")

    data = data.rename({
        'country': 'Country',
        'ISO_3digit': 'ISO3'
        },
        axis='columns'
    )

    data = data[[
        'Country',
        'ISO3',
    ]]

    for r in dataframe_to_rows(data, index=False, header=True):
        ws.append(r)

    ws['C1'] = "Scenario"
    ws['C2'] = "High"
    ws['C3'] = "Baseline"
    ws['C4'] = "Low"

    ws['D1'] = "Strategy"
    ws['D2'] = "4G" #"4G (Wireless)"
    # ws['D3'] = "4G (Fiber)"
    # ws['D4'] = "5G (Wireless)"
    # ws['D5'] = "5G (Fiber)"

    ws['E1'] = "Spectrum"
    ws['E2'] = "High"
    ws['E3'] = "Baseline"
    ws['E4'] = "Low"

    path = os.path.join(DATA_RAW, 'imf_gdp_2020_2030_real.csv')
    country_info = pd.read_csv(path, encoding = "ISO-8859-1")
    country_info.rename(columns={'isocode':'ISO3'}, inplace=True)

    country_info = country_info[[
        'ISO3',
        'ifscode',
        'region',
        'income'
    ]]

    my_list = [
        ('G', 'ISO3'),
        ('H', 'ifscode'),
        ('I', 'region'),
        ('J', 'income'),
    ]

    for item in my_list:
        col = item[0]
        metric = item[1]
        for idx, row in country_info.iterrows():

            if idx == 0:
                col_name_cell = '{}1'.format(col)
                ws[col_name_cell] = metric

            cell = '{}{}'.format(col, idx+2)
            ws[cell] = row[metric]
            ws.column_dimensions[col].width = 10

    return ws


def add_country_data(ws, metric):
    """
    Add the country data sheets.

    """
    ws.sheet_properties.tabColor = "666600"

    filename = 'all_pop_data_{}.csv'.format(metric)
    path = os.path.join(DATA_INTERMEDIATE, filename)
    data = pd.read_csv(path)

    data = data.rename({
        'GID_0': 'ISO3',
        }, axis='columns')

    lnth = len(data)+2

    for r in dataframe_to_rows(data, index=False, header=True):
        ws.append(r)

    ws.column_dimensions['B'].width = 30
    set_border(ws, 'A1:L{}'.format(len(data)+1), "thin", "000000")

    if metric == 'population':
        ws['M1'] = 'Population Sum'
        ws['N1'] = 'Income Group'
        ws['O1'] = 'Region'
        for i in range(2, lnth):
            cell = 'M{}'.format(i)
            ws[cell] = "=SUM(C{}:L{})".format(i,i)
            cell2 = 'N{}'.format(i)
            ws[cell2] = "=INDEX(Lookups!$M$2:$M$1611,MATCH(A{}, Lookups!$J$2:$J$1611,0))".format(i)
            cell3 = 'O{}'.format(i)
            ws[cell3] = "=INDEX(Lookups!$L$2:$L$1611,MATCH(A{}, Lookups!$J$2:$J$1611,0))".format(i)
        set_border(ws, 'M1:M{}'.format(len(data)+1), "thin", "000000")
        ws.column_dimensions['M'].width = 20
        set_border(ws, 'A1:O{}'.format(len(data)+1), "thin", "000000")

    if metric == 'area_km2':
        ws['M1'] = 'area_km2_sum'
        ws['N1'] = 'Income Group'
        ws['O1'] = 'Region'
        for i in range(2, lnth):
            cell = 'M{}'.format(i)
            ws[cell] = "=SUM(C{}:L{})".format(i,i)
            cell2 = 'N{}'.format(i)
            ws[cell2] = "=INDEX(Lookups!$M$2:$M$1611,MATCH(A{}, Lookups!$J$2:$J$1611,0))".format(i)
            cell3 = 'O{}'.format(i)
            ws[cell3] = "=INDEX(Lookups!$L$2:$L$1611,MATCH(A{}, Lookups!$J$2:$J$1611,0))".format(i)
        set_border(ws, 'M1:M{}'.format(len(data)+1), "thin", "000000")
        ws.column_dimensions['M'].width = 20
        set_border(ws, 'A1:O{}'.format(len(data)+1), "thin", "000000")

    return ws, lnth


def add_cum_pop(ws, cols):
    """
    Add the country data sheets.

    """
    ws.sheet_properties.tabColor = "666600"

    for col in cols:
        cell = "{}1".format(col)
        ws[cell] = "='Pop'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, 190):
            cell = "{}{}".format(col, i)
            ws[cell] = "='Pop'!{}".format(cell)

    for i in range(2, 190): #Decile
        cell = "C{}".format(i)
        part1 = "=IFERROR(INDEX(Pop!C$2:C$1611,MATCH(A{}, Pop!$A$2:$A$1611,0)),0)".format(i)
        ws[cell] = part1

    for idx, col in enumerate(cols[3:]):
        # print(idx, col, )
        for i in range(2, 190): #Decile
            cell = "{}{}".format(col, i)
            col_previous = cols[(2+idx):(3+idx)][0]
            part1 = "={}{}+IFERROR(INDEX(Pop!{}$2:{}$1611,MATCH(A{}, Pop!$A$2:$A$1611,0)),0)".format(col_previous, i, col,col,i)
            ws[cell] = part1

    set_border(ws, 'A1:L189', "thin", "000000")

    ws.column_dimensions['B'].width = 30

    return ws


def add_pop_growth(ws):
    """

    """
    ws.sheet_properties.tabColor = "666600"

    path = os.path.join(DATA_RAW, 'population_growth_rate_2020_2030.csv')
    p_growth = pd.read_csv(path, encoding = "ISO-8859-1")

    p_growth.rename(columns={
        'isocode':'ISO3',
        'country': 'country_name'
        }, inplace=True
    )

    p_growth = p_growth[[
        'ISO3','country_name','2020', '2021', '2022', '2023', '2024',
        '2025', '2026', '2027', '2028', '2029', '2030'
    ]]

    p_growth = p_growth.sort_values('ISO3')

    lnth = len(p_growth) + 2

    for r in dataframe_to_rows(p_growth, index=False, header=True):
        ws.append(r)

    ws['N1'] = 'Mean 10-Year Population Growth Rate (%)'
    for i in range(2,lnth):
        cell = 'N{}'.format(i)
        ws[cell] = "=SUM(C{}:M{})/11".format(i,i)

    ws.column_dimensions['N'].width = 40

    ws = format_numbers(ws, ['N'], (1, 200), 'Comma [0]', 1)

    set_border(ws, 'A1:N{}'.format(len(p_growth)+1), "thin", "000000")

    return ws, lnth


def add_context_data(ws, population, area, pop_density,
        pop_growth, users, data_demand):

    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    for col in cols:
        for i in range(1,5+1):
            cell = '{}{}'.format(col, i)
            if i == 1:
                ws[cell] = '=Pop!{}'.format(cell)
            else:
                if col == 'B':
                    ws[cell] = '=Countries!C{}'.format(i+6)
                if col == 'A':
                    ws[cell] = '=Countries!B{}'.format(i+6)
                if col not in ['A','B']:
                    ws[cell] = "=INDEX(Pop!{}2:{}1611,MATCH(B{}, Pop!A2:A1611,FALSE))/1e6".format(col, col, i)

    for i in range(1,5+1):
        for col in cols:
            if i == 1:
                old_cell = '{}{}'.format(col, i)
                new_cell = '{}{}'.format(col, i + 6)
                ws[new_cell] = '=Area!{}'.format(old_cell)
            else:
                new_cell = '{}{}'.format(col, i + 6)
                if col == 'B':
                    ws[new_cell] = '=Countries!C{}'.format(i+6)
                if col == 'A':
                    ws[new_cell] = '=Countries!B{}'.format(i+6)
                if col not in ['A','B']:
                    ws[new_cell] = "=INDEX(Area!{}2:{}1611,MATCH(B{}, Area!A2:A1611,FALSE))".format(col, col, i)

    for i in range(1,5+1):
        for col in cols:
            if i == 1:
                old_cell = '{}{}'.format(col, i)
                new_cell = '{}{}'.format(col, i + 12)
                ws[new_cell] = '=P_Density_km2!{}'.format(old_cell)
            else:
                new_cell = '{}{}'.format(col, i + 12)
                if col == 'B':
                    ws[new_cell] = '=Countries!C{}'.format(i+6)
                if col == 'A':
                    ws[new_cell] = '=Countries!B{}'.format(i+6)
                if col not in ['A','B']:
                    ws[new_cell] = "=INDEX(P_Density_km2!{}2:{}1611,MATCH(B{}, P_Density_km2!A2:A1611,FALSE))".format(col, col, i)

    for i in range(1,5+1):
        for col in cols:
            if i == 1:
                old_cell = '{}{}'.format(col, i)
                new_cell = '{}{}'.format(col, i + 18)
                ws[new_cell] = '=P_Growth!{}'.format(old_cell)
            else:
                new_cell = '{}{}'.format(col, i + 18)
                if col == 'B':
                    ws[new_cell] = '=Countries!C{}'.format(i+6)
                if col == 'A':
                    ws[new_cell] = '=Countries!B{}'.format(i+6)
                if col not in ['A','B']:
                    ws[new_cell] = "=INDEX(P_Growth!{}2:{}1611,MATCH(B{}, P_Growth!A2:A1611,FALSE))".format(col, col, i)

    ws['A31'] = "=TRANSPOSE(A1:L5)"
    ws['A44'] = "=TRANSPOSE(A7:L11)"
    ws['A57'] = "=TRANSPOSE(A13:L17)"
    ws['A70'] = "=TRANSPOSE(A19:L25)"

    ws.formula_attributes['A31'] = {'t': 'array', 'ref': "A31:E42"}
    ws.formula_attributes['A44'] = {'t': 'array', 'ref': "A44:E55"}
    ws.formula_attributes['A57'] = {'t': 'array', 'ref': "A57:E68"}
    ws.formula_attributes['A70'] = {'t': 'array', 'ref': "A70:E81"}

    return ws


def add_country_context(ws, data_sheet):
    """

    """
    ws.sheet_properties.tabColor = "92D050"

    ##Color white
    ws.sheet_view.showGridLines = False

    #Set blue and red border strips
    set_cell_color(ws, 'A1:AZ1', "004C97")
    set_cell_color(ws, 'A2:AZ2', "C00000")

    ws = decile_bar_chart(ws, "Context_Data!$B$32:$E$42", "Context_Data!$A$33:$A$42", "Population by Density Decile", 'Population Density Deciles\n(10 represents the first 10 percent of the densest statistical areas)', 'Population (Millions)', "B4")
    ws = decile_bar_chart(ws, "Context_Data!$B$45:$E$55", "Context_Data!$A$46:$A$55", "Area by Density Decile", 'Population Density Deciles\n(10 represents the first 10 percent of the densest statistical areas)', 'Area (Km^2)', "L4")
    ws = decile_bar_chart(ws, "Context_Data!$B$58:$E$68", "Context_Data!$A$59:$A$68", "Population Density by Density Decile", 'Population Density Deciles\n(10 represents the first 10 percent of the densest statistical areas)', 'Population Density (Km^2)', "B20")
    ws = decile_bar_chart(ws, "Context_Data!$B$71:$E$81", "Context_Data!$A$72:$A$81", "Population Growth Forecast", 'Year','Growth Rate (%)', "L20")

    return ws


def decile_bar_chart(ws, data, categories, title, x_axis, y_axis, loc):

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = title
    chart1.y_axis.title = y_axis
    chart1.x_axis.title = x_axis

    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(categories)
    chart1.shape = 4
    # chart1.dataLabels = DataLabelList()
    # chart1.dataLabels.showVal = True
    ws.add_chart(chart1, loc)

    return ws


def add_country_costs(ws, data_sheet):
    """

    """
    ws.sheet_properties.tabColor = "92D050"

    ##Color white
    ws.sheet_view.showGridLines = False

    #Set blue and red border strips
    set_cell_color(ws, 'A1:AZ1', "004C97")
    set_cell_color(ws, 'A2:AZ2', "C00000")

    ws = bar_chart(ws, "Estimates!$C$34:$C$38", "Estimates!$B$35:$B$38", "Total Cost by Country", 'Cost ($Bn)', "B4")
    ws = bar_chart(ws, "Estimates!$F$34:$F$38", "Estimates!$E$35:$E$38", "Mean Annual 10-Year GDP by Country",'GDP ($Bn)', "L4")
    ws = bar_chart(ws, "Estimates!$I$34:$I$38", "Estimates!$H$35:$H$38", "Initial Investment by Country",'Cost ($Bn)', "B20")
    ws = bar_chart(ws, "Estimates!$C$42:$C$46", "Estimates!$B$43:$B$46", "GDP Share by Country",'Percent of GDP (%)', "L20")

    return


def add_income_group_costs(ws, data_sheet):
    """

    """
    ws.sheet_properties.tabColor = "92D050"

    ##Color white
    ws.sheet_view.showGridLines = False

    #Set blue and red border strips
    set_cell_color(ws, 'A1:AZ1', "004C97")
    set_cell_color(ws, 'A2:AZ2', "C00000")

    ws = bar_chart(ws, "Estimates!$C$50:$C$53", "Estimates!$B$51:$B$53", "Total Cost by Income Group", 'Cost ($Bn)', "B4")
    ws = bar_chart(ws, "Estimates!$F$50:$F$53", "Estimates!$E$51:$E$53", "Mean Annual 10-Year GDP by Income Group",'GDP ($Bn)', "L4")
    ws = bar_chart(ws, "Estimates!$I$50:$I$53", "Estimates!$H$51:$H$53", "Initial Investment by Income Group",'Cost ($Bn)', "B20")
    ws = bar_chart(ws, "Estimates!$C$57:$C$60", "Estimates!$B$58:$B$60", "GDP Share by Income Group",'Percent of GDP (%)', "L20")

    return ws


def add_regional_group_costs(ws, data_sheet):
    """

    """
    ws.sheet_properties.tabColor = "92D050"

    ##Color white
    ws.sheet_view.showGridLines = False

    #Set blue and red border strips
    set_cell_color(ws, 'A1:AZ1', "004C97")
    set_cell_color(ws, 'A2:AZ2', "C00000")

    ws = bar_chart(ws, "Estimates!$C$64:$C$71", "Estimates!$B$65:$B$71", "Total Cost by Region", 'Cost ($Bn)', "B4")
    ws = bar_chart(ws, "Estimates!$F$64:$F$71", "Estimates!$E$65:$E$71", "Mean Annual 10-Year GDP by Region",'GDP ($Bn)', "L4")
    ws = bar_chart(ws, "Estimates!$I$64:$I$71", "Estimates!$H$65:$H$71", "Initial Investment by Region",'Cost ($Bn)', "B20")
    ws = bar_chart(ws, "Estimates!$C$75:$C$82", "Estimates!$B$76:$B$82", "GDP Share by Region",'Percent of GDP (%)', "L20")

    return ws


def bar_chart(ws, data, categories, title, y_axis, loc):

    chart1 = BarChart()
    chart1.type = "bar"
    chart1.title = title
    chart1.style = 10
    chart1.y_axis.title = y_axis
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(categories)
    chart1.shape = 4
    chart1.dataLabels = DataLabelList()
    chart1.dataLabels.showVal = True
    chart1.legend = None
    ws.add_chart(chart1, loc)

    return ws


def add_users(ws, cols, lnth):
    """
    Calculate the total number of users.

    """
    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='P_Density_km2'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='P_Density_km2'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth): #Decile
            cell = "{}{}".format(col, i)
            part1 = "='P_Density_km2'!{}".format(cell) #
            part2 = "*(POWER(1+(INDEX(P_Growth!$M$2:$M$1611,MATCH(A2, P_Growth!$A$2:$A$1611,0))/100), Settings!$C$9-Settings!$C$8))"
            part3 = "*(Settings!$C$15/100)*(Settings!$C$16/100)*(Settings!$C$17/100)"
            ws[cell] = part1 + part2 + part3

    set_border(ws, 'A1:L{}'.format(lnth-1), "thin", "000000")
    ws.column_dimensions['B'].width = 30

    return ws


def add_data_demand(ws, cols, lnth):
    """
    Calculate the total data demand.

    """
    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='Users_km2'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='Users_km2'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth): #Decile
            cell = "{}{}".format(col, i)
            part1 = "=(Users_km2!{}*MAX(Settings!$C$18,".format(cell)
            part2 = "(Settings!$C$19*1024*8*(1/30)*(15/100)*(1/3600))))"
            ws[cell] = part1 + part2

    set_border(ws, 'A1:L{}'.format(lnth-1), "thin", "000000")
    ws.column_dimensions['B'].width = 30

    return ws


def add_lookups_sheet(ws):
    """

    """
    ws.sheet_properties.tabColor = "66ffcc"

    ##Spectrum Portfolio Box
    ws.merge_cells('A1:B1')
    ws['A1'] = "Spectrum Portfolio"

    ws['A2'] = 'Scenario'
    ws['A3'] = 'High'
    ws['A4'] = 'Baseline'
    ws['A5'] = 'Low'

    ws['B2'] = 'Bandwidth (MHz)'
    ws['B3'] = 30
    ws['B4'] = 20
    ws['B5'] = 10

    set_border(ws, 'A1:B5', "thin", "000000")
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12

    ##Cost Information
    set_border(ws, 'A7:C13', "thin", "000000")
    ws.merge_cells('A7:C7')
    ws['A7'] = "Equipment Costs"

    ws['A8'] = 'Asset'
    ws['A9'] = 'RAN'
    ws['A10'] = 'Backhaul'
    ws['A11'] = 'Towers'
    ws['A12'] = 'Labor'
    ws['A13'] = 'Power System'

    ws['B8'] = 'Cost ($)'
    ws['B9'] = 30000
    ws['B10'] = 20000
    ws['B11'] = 15000
    ws['B12'] = 15000
    ws['B13'] = 2000

    ws['C8'] = 'Unit'
    ws['C9'] = "Per Tower"
    ws['C10'] = "Per Link"
    ws['C11'] = "Per Tower"
    ws['C12'] = "Per Tower"
    ws['C13'] = "Per Tower"

    ##Density Lookup Table
    ws.merge_cells('E1:H1')
    ws['E1'] = "Density Lookup Table"

    filename = 'capacity_lut_by_frequency.csv'
    path = os.path.join(DATA_INTERMEDIATE, 'luts', filename)
    lookup = pd.read_csv(path)
    # lookup.loc[lookup['frequency_GHz'] == '0.8']
    lookup = lookup[['sites_per_km2', 'capacity_mbps_km2']]
    lookup = lookup[lookup['capacity_mbps_km2'] != 0].reset_index()
    df_length = len(lookup)
    lookup = lookup.sort_values('sites_per_km2')

    my_list = [
        ('E', 'sites_per_km2'),
        ('F', 'capacity_mbps_km2'),
    ]
    for item in my_list:
        col = item[0]
        metric = item[1]
        for idx, row in lookup.iterrows():

            if idx == 0:
                col_name_cell = '{}2'.format(col)
                ws[col_name_cell] = metric

            cell = '{}{}'.format(col, idx+3)
            ws[cell] = row[metric]
            ws.column_dimensions[col].width = 18

    ws['G2'] = 'w_existing_spectrum'
    for i in range(3,250):
        col = 'G{}'.format(i)
        ws[col] = '=F{}*(VLOOKUP(Settings!$C$24, Lookups!$A$3:$B$5, 2, 0)/10)'.format(i)
    ws.column_dimensions['G'].width = 20

    ws['H2'] = 'w_future_spectrum'
    for i in range(3,250):
        col = 'H{}'.format(i)
        ws[col] = '=F{}*(VLOOKUP(Settings!$C$25, Lookups!$A$3:$B$5, 2, 0)/10)'.format(i)
    ws.column_dimensions['H'].width = 20

    set_border(ws, 'E1:H{}'.format(df_length+2), "thin", "000000")

    ###MNO quantities dataset
    filename = 'mno_quantities.csv'
    path = os.path.join(DATA_RAW, filename)
    lookup = pd.read_csv(path, encoding = "ISO-8859-1")#, header=True
    lookup = lookup[['iso3', 'quantity', 'region', 'income']]
    lookup = lookup.sort_values('iso3')

    for idx, item in lookup.iterrows():

        if idx == 0:
            ws['J1'] = 'iso3'
            ws['K1'] = 'quantity'
            ws['L1'] = 'region'
            ws['M1'] = 'income'

        cell1 = 'J{}'.format(idx+2)
        ws[cell1] = item['iso3']
        cell2 = 'K{}'.format(idx+2)
        ws[cell2] = item['quantity']
        cell3 = 'L{}'.format(idx+2)
        ws[cell3] = item['region']
        cell4 = 'M{}'.format(idx+2)
        ws[cell4] = item['income']

    set_border(ws, 'J1:M{}'.format(len(lookup)+1), "thin", "000000")

    ws.column_dimensions['J'].width = 8
    ws.column_dimensions['K'].width = 8
    ws.column_dimensions['L'].width = 25
    ws.column_dimensions['M'].width = 25

    return ws


def add_coverage_sheet(ws, cols):
    """

    """
    path = os.path.join(DATA_RAW, 'gsma_3g_coverage.csv')
    coverage = pd.read_csv(path, encoding = "ISO-8859-1")
    coverage['coverage'] = coverage['coverage'] * 100
    coverage = coverage.sort_values('ISO3')
    coverage = coverage.dropna()
    coverage = coverage[['ISO3', 'country_name', 'coverage']]#.reset_index()

    path = os.path.join(DATA_RAW, 'coverage_4g.csv')
    coverage_4g = pd.read_csv(path, encoding = "ISO-8859-1")
    coverage_4g = coverage_4g[['ISO3', 'coverage_4G']]#.reset_index()
    coverage_4g = coverage_4g.sort_values('ISO3')#.reset_index()
    coverage_4g = coverage_4g.dropna()

    path = os.path.join(DATA_RAW, 'site_counts', 'site_counts.csv')
    sites = pd.read_csv(path, encoding = "ISO-8859-1")
    sites = sites[['ISO3', 'sites']]

    coverage = coverage.merge(coverage_4g, left_on='ISO3', right_on='ISO3', right_index=False)

    coverage = coverage.merge(sites, left_on='ISO3', right_on='ISO3')

    for r in dataframe_to_rows(coverage, index=False, header=True):
        ws.append(r)

    lnth = len(coverage)+2

    ws['F1'] = 'Population'
    for i in range(2, lnth):
        cell = "F{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Pop!$M$2:$M$1611,MATCH(A{}, Pop!$A$2:$A$1611,0)),"")".format(i)

    ws['G1'] = 'Sites per covered population'
    for i in range(2, lnth):
        cell = "G{}".format(i)
        ws[cell] = '=IFERROR(F{}*(C{}/100)/E{}, "-")'.format(i,i,i)

    ws['H1'] = 'Covered population'
    for i in range(2, lnth):
        cell = "H{}".format(i)
        ws[cell] = '=IFERROR(F{}*(C{}/100), "-")'.format(i,i)

    ws['I1'] = 'Covered pop w/4G'
    for i in range(2, lnth):
        cell = "I{}".format(i)
        ws[cell] = '=IFERROR(F{}*(D{}/100), "-")'.format(i,i)

    set_border(ws, 'A1:I{}'.format(len(coverage)+1), "thin", "000000")

    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15

    return ws#, lnth


def add_towers_sheet(ws, cols, lnth):
    """

    """
    for col in cols:
        cell = "{}1".format(col)
        ws[cell] = "='Data_km2'!{}".format(cell)

    for col in cols[:2]:
        for i in range(1, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='Data_km2'!{}".format(cell)

    for i in range(2, lnth): #Decile
        cell = "C{}".format(i)
        part1 = "=IFERROR(INDEX(Pop!$C$2:$C$1611,MATCH(A{}, Pop!$A$2:$A$1611,0)) /".format(i)
        part2 = "INDEX(Coverage!$G$2:$G$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),0)".format(i)
        ws[cell] = part1 + part2 #+ part3

    for i in range(2, lnth):
        cell = "D{}".format(i)
        part1 = "=IFERROR(IF(SUM(C{})<INDEX(Coverage!$E$2:$E$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),".format(i,i)
        part2 = "INDEX(Pop!$D$2:$D$1611,MATCH(A{}, Pop!$A$2:$A$1611,0))/".format(i)
        part3 = "INDEX(Coverage!$G$2:$G$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),0),0)".format(i)
        ws[cell] = part1 + part2 + part3

    for i in range(2, lnth):
        cell = "E{}".format(i)
        part1 = "=IFERROR(IF(SUM(C{}:D{})<INDEX(Coverage!$E$2:$E$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),".format(i,i,i)
        part2 = "INDEX(Pop!$E$2:$E$1611,MATCH(A{}, Pop!$A$2:$A$1611,0))/".format(i)
        part3 = "INDEX(Coverage!$G$2:$G$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),0),0)".format(i)
        ws[cell] = part1 + part2 + part3

    for i in range(2, lnth):
        cell = "F{}".format(i)
        part1 = "=IFERROR(IF(SUM(C{}:E{})<INDEX(Coverage!$E$2:$E$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),".format(i,i,i)
        part2 = "INDEX(Pop!$F$2:$F$1611,MATCH(A{}, Pop!$A$2:$A$1611,0))/".format(i)
        part3 = "INDEX(Coverage!$G$2:$G$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),0),0)".format(i)
        ws[cell] = part1 + part2 + part3

    for i in range(2, lnth):
        cell = "G{}".format(i)
        part1 = "=IFERROR(IF(SUM(C{}:F{})<INDEX(Coverage!$E$2:$E$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),".format(i,i,i)
        part2 = "INDEX(Pop!$G$2:$G$1611,MATCH(A{}, Pop!$A$2:$A$1611,0))/".format(i)
        part3 = "INDEX(Coverage!$G$2:$G$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),0),0)".format(i)
        ws[cell] = part1 + part2 + part3

    for i in range(2, lnth):
        cell = "H{}".format(i)
        part1 = "=IFERROR(IF(SUM(C{}:G{})<INDEX(Coverage!$E$2:$E$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),".format(i,i,i)
        part2 = "INDEX(Pop!$H$2:$H$1611,MATCH(A{}, Pop!$A$2:$A$1611,0))/".format(i)
        part3 = "INDEX(Coverage!$G$2:$G$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),0),0)".format(i)
        ws[cell] = part1 + part2 + part3

    for i in range(2, lnth):
        cell = "I{}".format(i)
        part1 = "=IFERROR(IF(SUM(C{}:H{})<INDEX(Coverage!$E$2:$E$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),".format(i,i,i)
        part2 = "INDEX(Pop!$I$2:$I$1611,MATCH(A{}, Pop!$A$2:$A$1611,0))/".format(i)
        part3 = "INDEX(Coverage!$G$2:$G$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),0),0)".format(i)
        ws[cell] = part1 + part2 + part3

    for i in range(2, lnth):
        cell = "J{}".format(i)
        part1 = "=IFERROR(IF(SUM(C{}:I{})<INDEX(Coverage!$E$2:$E$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),".format(i,i,i)
        part2 = "INDEX(Pop!$J$2:$J$1611,MATCH(A{}, Pop!$A$2:$A$1611,0))/".format(i)
        part3 = "INDEX(Coverage!$G$2:$G$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),0),0)".format(i)
        ws[cell] = part1 + part2 + part3

    for i in range(2, lnth):
        cell = "K{}".format(i)
        part1 = "=IFERROR(IF(SUM(C{}:J{})<INDEX(Coverage!$E$2:$E$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),".format(i,i,i)
        part2 = "INDEX(Pop!$K$2:$K$1611,MATCH(A{}, Pop!$A$2:$A$1611,0))/".format(i)
        part3 = "INDEX(Coverage!$G$2:$G$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),0),0)".format(i)
        ws[cell] = part1 + part2 + part3

    for i in range(2, lnth):
        cell = "L{}".format(i)
        part1 = "=IFERROR(IF(SUM(C{}:K{})<INDEX(Coverage!$E$2:$E$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),".format(i,i,i)
        part2 = "INDEX(Pop!$L$2:$L$1611,MATCH(A{}, Pop!$A$2:$A$1611,0))/".format(i)
        part3 = "INDEX(Coverage!$G$2:$G$1611,MATCH(A{}, Coverage!$A$2:$A$1611,0)),0),0)".format(i)
        ws[cell] = part1 + part2 + part3

    set_border(ws, 'A1:L{}'.format(lnth-1), "thin", "000000")

    return ws


def add_towers_4G_sheet(ws, cols, lnth):
    """

    """
    for col in cols:
        cell = "{}1".format(col)
        ws[cell] = "=Towers!{}".format(cell)

    for col in cols[:2]:
        for i in range(1, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "=Towers!{}".format(cell)

    # for col in cols[:3]:
    for i in range(2, lnth):
        cell = "C{}".format(i)
        part1 = "=IFERROR(IF(Pop_C!C{}<IFERROR(INDEX(Coverage!$I$2:$I$1611,MATCH(A{},Coverage!$A$2:$A$1611,0)),0),Towers!C{},".format(i,i,i)
        part2 = "(IFERROR(INDEX(Coverage!$I$2:$I$1611,MATCH(A5,Coverage!$A$2:$A$1611,0)),0)/Pop_C!C{}*Towers!C{})),0)".format(i,i)
        ws[cell] = part1 + part2

    for idx, col in enumerate(cols[3:]):
        # print(idx, col, )
        for i in range(2, lnth): #Decile
            cell = "{}{}".format(col, i)
            # col_previous = cols[(2+idx):(3+idx)][0]
            # part1 = "={}{}+IFERROR(INDEX(Pop!{}$2:{}$1611,MATCH(A{}, Pop!$A$2:$A$1611,0)),0)".format(col_previous, i, col,col,i)
            part1 = "=IFERROR(IF(Pop_C!{}<IFERROR(INDEX(Coverage!$I$2:$I$1611,MATCH(A{},Coverage!$A$2:$A$1611,0)),0),Towers!{},0),0)".format(cell,i,cell)
            ws[cell] = part1
            ws.formula_attributes[cell] = {'t': 'array', 'ref': "{}:{}".format(cell, cell)}

    set_border(ws, 'A1:L{}'.format(lnth-1), "thin", "000000")

    return ws


def add_towers_4G_km2_sheet(ws, cols, lnth):
    """

    """
    for col in cols:
        cell = "{}1".format(col)
        ws[cell] = "=Towers!{}".format(cell)

    for col in cols[:2]:
        for i in range(1, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "=Towers!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "=IFERROR(Towers!{}/Area!{},0)".format(cell,cell)

    set_border(ws, 'A1:L{}'.format(lnth-1), "thin", "000000")

    return ws


def add_towers_mno_sheet(ws, cols, lnth):
    """

    """
    for col in cols:
        cell = "{}1".format(col)
        ws[cell] = "=Towers!{}".format(cell)

    for col in cols[:2]:
        for i in range(1, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "=Towers!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "=IFERROR(Towers!{}*(Settings!C16/100),0)".format(cell)

    set_border(ws, 'A1:L{}'.format(lnth-1), "thin", "000000")

    return ws


def add_towers_4G_mno_sheet(ws, cols, lnth):
    """

    """
    for col in cols:
        cell = "{}1".format(col)
        ws[cell] = "=Towers_4G!{}".format(cell)

    for col in cols[:2]:
        for i in range(1, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "=Towers_4G!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "=IFERROR(Towers_4G!{}*(Settings!C16/100),0)".format(cell)

    set_border(ws, 'A1:L{}'.format(lnth-1), "thin", "000000")

    return ws


def add_towers_non_4G_mno_sheet(ws, cols, lnth):
    """

    """
    for col in cols:
        cell = "{}1".format(col)
        ws[cell] = "=Towers_4G!{}".format(cell)

    for col in cols[:2]:
        for i in range(1, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "=Towers_4G!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "=IF(Towers_MNO!{}-Towers_4G_MNO!{}<0,0,Towers_MNO!{}-Towers_4G_MNO!{})".format(cell,cell,cell,cell)

    set_border(ws, 'A1:L{}'.format(lnth-1), "thin", "000000")

    return ws


def add_capacity_sheet(ws, cols, lnth):
    """

    """
    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='Data_km2'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='Data_km2'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth): #Total_Sites Density
            cell = "{}{}".format(col, i)
            part1 = '=IFERROR(MAX(IF(Lookups!$E$3:$E$250<Towers_4G_km2!{}'.format(cell)
            part2 = '*(Settings!$C$16/100),Lookups!$G$3:$G$250)),"-")'
            ws[cell] = part1 + part2
            ws.formula_attributes[cell] = {'t': 'array', 'ref': "{}:{}".format(cell, cell)}

    set_border(ws, 'A1:L{}'.format(lnth-1), "thin", "000000")

    return ws


def add_sites_sheet(ws, cols, lnth):
    """

    """
    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='Capacity_km2_MNO'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='Capacity_km2_MNO'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            part1 = "=MIN(IF('Lookups'!$H$3:$H$250>'Data_km2'!{}".format(cell)
            part2 = ",'Lookups'!$E$3:$E$250))*Area!{}".format(cell)
            ws[cell] = part1 + part2
            ws.formula_attributes[cell] = {'t': 'array', 'ref': "{}:{}".format(cell, cell)}

    columns = ['C','D','E','F','G','H','I','J','K','L']
    ws = format_numbers(ws, columns, (1, 200), 'Comma [0]', 0)

    set_border(ws, 'A1:L{}'.format(lnth-1), "thin", "000000")

    return ws


def add_sites_km2_sheet(ws, cols, lnth):
    """

    """
    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='Capacity_km2_MNO'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='Capacity_km2_MNO'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            part1 = "=MIN(IF('Lookups'!$H$3:$H$250>'Data_km2'!{}".format(cell)
            part2 = ",'Lookups'!$E$3:$E$250))"#*Area!{}".format(cell)
            ws[cell] = part1 + part2
            ws.formula_attributes[cell] = {'t': 'array', 'ref': "{}:{}".format(cell, cell)}

    columns = ['C','D','E','F','G','H','I','J','K','L']
    ws = format_numbers(ws, columns, (1, 200), 'Comma [0]', 5)

    set_border(ws, 'A1:L{}'.format(lnth-1), "thin", "000000")

    return ws


def add_new_sites_sheet(ws, cols, lnth):
    """

    """
    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='Total_Sites_MNO'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='Total_Sites_MNO'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            part1 = "=IF(P_Density_km2!{}>Settings!$C$26,".format(cell)
            part2 = "IF(Towers_4G_MNO!{}<Total_Sites_MNO!{},".format(cell, cell)
            part3 = "(Total_Sites_MNO!{}-Towers_4G_MNO!{}),0),".format(cell, cell)
            part4 = '"-")'
            ws[cell] = part1 + part2 + part3 + part4

    columns = ['C','D','E','F','G','H','I','J','K','L']
    ws = format_numbers(ws, columns, (1, 200), 'Comma [0]', 0)

    set_border(ws, 'A1:L{}'.format(lnth-1), "thin", "000000")

    return ws


def add_upgrades(ws, cols, lnth):
    """

    """
    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            part1 = "=IFERROR(IF(Towers_non_4G_MNO!{}>0,IF(Towers_non_4G_MNO!{}>".format(cell,cell)
            part2 = "New_4G_Sites!{},New_4G_Sites!{},New_4G_Sites!{}-Towers_non_4G_MNO!{}),0),0)".format(cell,cell,cell,cell)
            ws[cell] = part1 + part2 #+ part3 + part4

    columns = ['C','D','E','F','G','H','I','J','K','L']
    ws = format_numbers(ws, columns, (1, 200), 'Comma [0]', 0)

    set_border(ws, 'A1:L{}'.format(lnth-1), "thin", "000000")

    return ws


def add_new_builds(ws, cols, lnth):
    """

    """
    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            part1 = "=IFERROR(New_4G_Sites!{}-Upgrades!{},0)".format(cell,cell)
            ws[cell] = part1

    columns = ['C','D','E','F','G','H','I','J','K','L']
    ws = format_numbers(ws, columns, (1, 200), 'Comma [0]', 0)

    set_border(ws, 'A1:L{}'.format(lnth-1), "thin", "000000")

    return ws


def add_site_users_sheet(ws, cols, lnth):
    """

    """
    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='Total_Sites_MNO'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='Total_Sites_MNO'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            part1 = "=IFERROR((Users_km2!{}*Area!{})/Total_Sites_MNO!{},0)".format(cell,cell,cell)
            ws[cell] = part1
            ws.formula_attributes[cell] = {'t': 'array', 'ref': "{}:{}".format(cell, cell)}

    columns = ['C','D','E','F','G','H','I','J','K','L']
    ws = format_numbers(ws, columns, (1, 200), 'Comma [0]', 0)

    set_border(ws, 'A1:L{}'.format(lnth-1), "thin", "000000")

    return ws


def add_site_costs(ws, cols, lnth):
    """

    """
    ws.sheet_properties.tabColor = "9966ff"

    #Deciles
    # set_border(ws, 'A1:J11', "thin", "000000")

    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            # =IFERROR('New_4G_Sites'!K2*VLOOKUP(Lookups!$A$9, Lookups!$A$9:Lookups!$B$24, 2, FALSE),"-")
            part1 = '=IFERROR(New_4G_Sites!{}*VLOOKUP(Lookups!$A$9, Lookups!$A$9:Lookups!$B$24, 2, FALSE),"-")'.format(cell)
            # part2 = "+((SQRT((1/('Total_Sites_MNO'!{}/Area!{}))/2)*1000)*'New_4G_Sites'!{}".format(cell, cell, cell)
            # part3 = "*VLOOKUP('Lookups'!$A$10, 'Lookups'!$A$9:'Lookups'!$B$24, 2, FALSE))"
            # part4 = "+'New_4G_Sites'!{}*VLOOKUP('Lookups'!$A$11, 'Lookups'!$A$9:'Lookups'!$B$24, 2, FALSE)".format(cell)
            # part5 = ',"-")'
            ws[cell] = part1 #+ part2 + part3 + part4 + part5

    ws['M1'] = 'Income Group'
    for i in range(2,lnth):
        cell = "M{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$J$2:$J$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws['N1'] = 'Region'
    for i in range(2,lnth):
        cell = "N{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$I$2:$I$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 35
    ws.column_dimensions['P'].width = 45

    ws = format_numbers(ws, ['N'], (1,200), 'Comma [0]', 0)
    ws = format_numbers(ws, ['M'], (1,200), 'Comma [0]', 1)

    set_border(ws, 'A1:N{}'.format(lnth-1), "thin", "000000")

    return ws


def add_bh_costs(ws, cols, lnth):
    """

    """
    ws.sheet_properties.tabColor = "9966ff"

    #Deciles
    # set_border(ws, 'A1:J11', "thin", "000000")

    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            # part1 = '=IFERROR(New_4G_Sites!{}*(SQRT((1/(Total_Sites_km2!{}/Area!{}))/2)*1000)*VLOOKUP(Lookups!$A$10, Lookups!$A$9:Lookups!$B$24, 2, FALSE),"-")'.format(cell, cell, cell)
            # part1 = "=New_4G_Sites!{}*VLOOKUP(Lookups!$A$10,Lookups!$A$9:Lookups!$B$24,2,FALSE)".format(cell)
            # part1 = "=IFERROR(ROUND(SQRT((1/Total_Sites_km2!{})/2),0),0)".format(cell)
            part1 = "=IFERROR(ROUNDUP(SQRT((1/Total_Sites_km2!{})/2)/20,0)*New_4G_Sites!{}*VLOOKUP(Lookups!$A$10,Lookups!$A$9:Lookups!$B$24,2,FALSE),0)".format(cell,cell)
            ws[cell] = part1 #+ part2 + part3 + part4 + part5

    ws['M1'] = 'Income Group'
    for i in range(2,lnth):
        cell = "M{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$J$2:$J$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws['N1'] = 'Region'
    for i in range(2,lnth):
        cell = "N{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$I$2:$I$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 35
    ws.column_dimensions['P'].width = 45

    ws = format_numbers(ws, ['N'], (1,200), 'Comma [0]', 0)
    ws = format_numbers(ws, ['M'], (1,200), 'Comma [0]', 1)

    set_border(ws, 'A1:N{}'.format(lnth-1), "thin", "000000")

    return ws


def add_tower_costs(ws, cols, lnth):
    """

    """
    ws.sheet_properties.tabColor = "9966ff"

    #Deciles
    # set_border(ws, 'A1:J11', "thin", "000000")

    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            part1 = '=IFERROR(New_Builds!{}*VLOOKUP(Lookups!$A$11, Lookups!$A$9:Lookups!$B$24, 2, FALSE), "-")'.format(cell)
            ws[cell] = part1

    ws['M1'] = 'Income Group'
    for i in range(2,lnth):
        cell = "M{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$J$2:$J$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws['N1'] = 'Region'
    for i in range(2,lnth):
        cell = "N{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$I$2:$I$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 35
    ws.column_dimensions['P'].width = 45

    ws = format_numbers(ws, ['N'], (1,200), 'Comma [0]', 0)
    ws = format_numbers(ws, ['M'], (1,200), 'Comma [0]', 1)

    set_border(ws, 'A1:N{}'.format(lnth-1), "thin", "000000")

    return ws


def add_labor_costs(ws, cols, lnth):
    """

    """
    ws.sheet_properties.tabColor = "9966ff"

    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            # part1 = "=IFERROR('New_4G_Sites'!{}*VLOOKUP(Lookups!$A$9, Lookups!$A$9:Lookups!$B$24, 2, FALSE),0)".format(cell)
            # part1 = '=IFERROR(New_4G_Sites!{}*(SQRT((1/(Total_Sites_MNO!{}/Area!{}))/2)*1000),"-")'.format(cell, cell, cell)
            # part1 = "*VLOOKUP('Lookups'!$A$10, 'Lookups'!$A$9:'Lookups'!$B$24, 2, FALSE))"
            part1 = '=IFERROR(New_4G_Sites!{}*VLOOKUP(Lookups!$A$12, Lookups!$A$9:Lookups!$B$24, 2, FALSE), "-")'.format(cell)
            # part5 = ',"-")'
            ws[cell] = part1 #+ part2 + part3 + part4 + part5

    ws['M1'] = 'Income Group'
    for i in range(2,lnth):
        cell = "M{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$J$2:$J$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws['N1'] = 'Region'
    for i in range(2,lnth):
        cell = "N{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$I$2:$I$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 35
    ws.column_dimensions['P'].width = 45

    ws = format_numbers(ws, ['N'], (1,200), 'Comma [0]', 0)
    ws = format_numbers(ws, ['M'], (1,200), 'Comma [0]', 1)

    set_border(ws, 'A1:N{}'.format(lnth-1), "thin", "000000")

    return ws


def add_power_costs(ws, cols, lnth):
    """

    """
    ws.sheet_properties.tabColor = "9966ff"

    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            # part1 = "=IFERROR('New_4G_Sites'!{}*VLOOKUP(Lookups!$A$9, Lookups!$A$9:Lookups!$B$24, 2, FALSE),0)".format(cell)
            # part1 = '=IFERROR(New_4G_Sites!{}*(SQRT((1/(Total_Sites_MNO!{}/Area!{}))/2)*1000),"-")'.format(cell, cell, cell)
            # part1 = "*VLOOKUP('Lookups'!$A$10, 'Lookups'!$A$9:'Lookups'!$B$24, 2, FALSE))"
            part1 = '=IFERROR(New_4G_Sites!{}*VLOOKUP(Lookups!$A$13, Lookups!$A$9:Lookups!$B$24, 2, FALSE), "-")'.format(cell)
            # part5 = ',"-")'
            ws[cell] = part1 #+ part2 + part3 + part4 + part5

    ws['M1'] = 'Income Group'
    for i in range(2,lnth):
        cell = "M{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$J$2:$J$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws['N1'] = 'Region'
    for i in range(2,lnth):
        cell = "N{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$I$2:$I$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 35
    ws.column_dimensions['P'].width = 45

    ws = format_numbers(ws, ['N'], (1,200), 'Comma [0]', 0)
    ws = format_numbers(ws, ['M'], (1,200), 'Comma [0]', 1)

    set_border(ws, 'A1:N{}'.format(lnth-1), "thin", "000000")

    return ws


def add_site_opex(ws, cols, lnth):
    """

    """
    ws.sheet_properties.tabColor = "9966ff"

    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            part1 = '=IFERROR(((RAN_Capex!{}*0.1)*Settings!C16)/((1+Settings!C11)^Settings!C16), "-")'.format(cell)
            ws[cell] = part1 #+ part2 + part3 + part4 + part5

    ws['M1'] = 'Income Group'
    for i in range(2,lnth):
        cell = "M{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$J$2:$J$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws['N1'] = 'Region'
    for i in range(2,lnth):
        cell = "N{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$I$2:$I$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 35
    ws.column_dimensions['P'].width = 45

    ws = format_numbers(ws, ['N'], (1,200), 'Comma [0]', 0)
    ws = format_numbers(ws, ['M'], (1,200), 'Comma [0]', 1)

    set_border(ws, 'A1:N{}'.format(lnth-1), "thin", "000000")

    return ws


def add_bh_opex(ws, cols, lnth):
    """

    """
    ws.sheet_properties.tabColor = "9966ff"

    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            part1 = '=IFERROR(((BH_Capex!{}*0.1)*Settings!C16)/((1+Settings!C11)^Settings!C16), "-")'.format(cell)
            ws[cell] = part1 #+ part2 + part3 + part4 + part5

    ws['M1'] = 'Income Group'
    for i in range(2,lnth):
        cell = "M{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$J$2:$J$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws['N1'] = 'Region'
    for i in range(2,lnth):
        cell = "N{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$I$2:$I$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 35
    ws.column_dimensions['P'].width = 45

    ws = format_numbers(ws, ['N'], (1,200), 'Comma [0]', 0)
    ws = format_numbers(ws, ['M'], (1,200), 'Comma [0]', 1)

    set_border(ws, 'A1:N{}'.format(lnth-1), "thin", "000000")

    return ws


def add_tower_opex(ws, cols, lnth):
    """

    """
    ws.sheet_properties.tabColor = "9966ff"

    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            part1 = '=IFERROR(((Tower_Capex!{}*0.1)*Settings!C16)/((1+Settings!C11)^Settings!C16), "-")'.format(cell)
            ws[cell] = part1 #+ part2 + part3 + part4 + part5

    ws['M1'] = 'Income Group'
    for i in range(2,lnth):
        cell = "M{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$J$2:$J$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws['N1'] = 'Region'
    for i in range(2,lnth):
        cell = "N{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$I$2:$I$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 35
    ws.column_dimensions['P'].width = 45

    ws = format_numbers(ws, ['N'], (1,200), 'Comma [0]', 0)
    ws = format_numbers(ws, ['M'], (1,200), 'Comma [0]', 1)

    set_border(ws, 'A1:N{}'.format(lnth-1), "thin", "000000")

    return ws


def add_power_opex(ws, cols, lnth):
    """

    """
    ws.sheet_properties.tabColor = "9966ff"

    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            part1 = '=IFERROR(((Tower_Capex!{}*0.1)*Settings!C16)/((1+Settings!C11)^Settings!C16), "-")'.format(cell)
            ws[cell] = part1 #+ part2 + part3 + part4 + part5

    ws['M1'] = 'Income Group'
    for i in range(2,lnth):
        cell = "M{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$J$2:$J$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws['N1'] = 'Region'
    for i in range(2,lnth):
        cell = "N{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$I$2:$I$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 35
    ws.column_dimensions['P'].width = 45

    ws = format_numbers(ws, ['N'], (1,200), 'Comma [0]', 0)
    ws = format_numbers(ws, ['M'], (1,200), 'Comma [0]', 1)

    set_border(ws, 'A1:N{}'.format(lnth-1), "thin", "000000")

    return ws


def add_mno_costs(ws, cols, lnth):
    """

    """
    ws.sheet_properties.tabColor = "9966ff"

    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            part1 = '=IFERROR(RAN_Capex!{}+BH_Capex!{}+Tower_Capex!{}+Labor_Capex!{}+Power_Capex!{}+RAN_Opex!{}+BH_Opex!{}+Tower_Opex!{}+Power_Opex!{}, "-")'.format(cell,cell,cell,cell,cell,cell,cell,cell,cell)
            ws[cell] = part1

    ws['M1'] = 'MNO Cost ($)'
    for i in range(2,lnth):
        cell = "M{}".format(i)
        part1 = '=IFERROR(SUMIF((C{}:L{}), "<>n/a"), "-")'.format(i, i)
        line = part1
        ws[cell] = line

    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 35
    ws.column_dimensions['P'].width = 45

    ws = format_numbers(ws, ['N'], (1,200), 'Comma [0]', 0)
    ws = format_numbers(ws, ['M'], (1,200), 'Comma [0]', 1)

    set_border(ws, 'A1:N{}'.format(lnth-1), "thin", "000000")

    return ws


def add_total_costs(ws, cols, lnth):
    """

    """
    ws.sheet_properties.tabColor = "9966ff"

    for col in cols:
            cell = "{}1".format(col)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[:2]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            ws[cell] = "='New_4G_Sites'!{}".format(cell)

    for col in cols[2:]:
        for i in range(2, lnth):
            cell = "{}{}".format(col, i)
            part1 = '=IFERROR(MNO_Costs!{}*(100/Settings!C16), "-")'.format(cell)
            ws[cell] = part1

    ws['M1'] = 'Total Cost ($)'
    for i in range(2,lnth):
        cell = "M{}".format(i)
        part1 = '=IFERROR(SUMIF((C{}:L{}), "<>n/a"), "-")'.format(i, i)
        line = part1
        ws[cell] = line

    ws['N1'] = 'Cost Per Pop ($)'
    for i in range(2,lnth):
        cell = "N{}".format(i)
        ws[cell] = "=(M{})/Pop!M{}".format(i, i)

    ws['O1'] = 'Income Group'
    for i in range(2,lnth):
        cell = "O{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$J$2:$J$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws['P1'] = 'Region'
    for i in range(2,lnth):
        cell = "P{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$I$2:$I$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 35
    ws.column_dimensions['P'].width = 45

    ws = format_numbers(ws, ['N'], (1,200), 'Comma [0]', 0)
    ws = format_numbers(ws, ['M'], (1,200), 'Comma [0]', 1)

    set_border(ws, 'A1:N{}'.format(lnth-1), "thin", "000000")

    return ws


def add_gdp_sheet(ws):
    """

    """
    ws.sheet_properties.tabColor = "ffff33"

    path = os.path.join(DATA_RAW, 'imf_gdp_2020_2030_real.csv')
    gdp = pd.read_csv(path, encoding = "ISO-8859-1")

    gdp.rename(columns={'isocode':'ISO3'}, inplace=True)

    for i in range(2020, 2031):
        col = "GDP{}".format(i)
        if col in gdp.columns:
            gdp.rename(columns={col:i}, inplace=True)

    gdp = gdp[['ISO3',2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030]]

    gdp = gdp.sort_values('ISO3')

    lnth = len(gdp) + 2

    for r in dataframe_to_rows(gdp, index=False, header=True):
        ws.append(r)

    ws['L1'] = 'Mean 10-Year GDP ($Bn)'
    for i in range(2,lnth):
        cell = 'L{}'.format(i)
        ws[cell] = "=((SUM(B{}:K{})*1e9)/10)/1e9".format(i,i)

    ws['M1'] = 'GDP Growth Rate (%)'
    for i in range(2,lnth):
        cell = 'M{}'.format(i)
        ws[cell] = "=IFERROR((K{}-B{})/B{},"")".format(i,i,i)
    ws = format_numbers(ws, ['M'], (2,len(gdp)+1), 'Percent', 1)

    ws['N1'] = 'Income Group'
    for i in range(2,lnth):
        cell = "N{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$J$2:$J$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws['O1'] = 'Region'
    for i in range(2,lnth):
        cell = "O{}".format(i)
        ws[cell] = "=IFERROR(INDEX(Options!$I$2:$I$1611,MATCH(A{}, Options!$G$2:$G$1611,0)), "")".format(i)

    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 35
    ws.column_dimensions['O'].width = 45

    ws = format_numbers(ws, ['L'], (1, 200), 'Comma [0]', 1)

    set_border(ws, 'A1:O{}'.format(len(gdp)+1), "thin", "000000")

    return ws


if __name__ == "__main__":

    generate_workbook()

    # try:
    extract_data()
    # except:
    #     print('Problem extracting results')
