from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference

wb = Workbook(write_only=True)
ws = wb.create_sheet()

rows = [
    ('Number', 'Batch 1'),#, 'Batch 2'),
    (2, 10),# 30),
    (3, 40),# 60),
    (4, 50),# 70),
    (5, 20),# 10),
    (6, 10),# 40),
    (7, 50),# 30),
]

for row in rows:
    ws.append(row)

chart1 = BarChart()
chart1.type = "bar"
chart1.title = "Horizontal Bar Chart"
chart1.style = 10
chart1.y_axis.title = 'Test number'
chart1.x_axis.title = 'Sample length (mm)'

data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=2)
cats = Reference(ws, min_col=1, min_row=2, max_row=7)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
ws.add_chart(chart1, "A10")

# from copy import deepcopy

# chart2 = deepcopy(chart1)

# chart1 = BarChart()
# chart1.type = "col"
# chart1.style = 10
# chart1.title = "Bar Chart"
# chart1.y_axis.title = 'Test number'
# chart1.x_axis.title = 'Sample length (mm)'

# data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=2)
# cats = Reference(ws, min_col=1, min_row=2, max_row=7)
# chart1.add_data(data, titles_from_data=True)
# chart1.set_categories(cats)
# chart1.shape = 4
# ws.add_chart(chart1, "A10")

# chart2.style = 11
# chart2.type = "bar"
# chart2.title = "Horizontal Bar Chart"

# ws.add_chart(chart2, "J10")


# chart3 = deepcopy(chart1)
# chart3.type = "col"
# chart3.style = 12
# chart3.grouping = "stacked"
# chart3.overlap = 100
# chart3.title = 'Stacked Chart'

# ws.add_chart(chart3, "A27")


# chart4 = deepcopy(chart1)
# chart4.type = "bar"
# chart4.style = 13
# chart4.grouping = "percentStacked"
# chart4.overlap = 100
# chart4.title = 'Percent Stacked Chart'

# ws.add_chart(chart4, "G27")

wb.save("vis_test.xlsx")

# from openpyxl import Workbook
# from openpyxl.chart import BarChart, Reference, Series

# wb = Workbook()
# wb.create_sheet('Data')

# ws = wb['Data']
# for i in range(10):
#     ws.append([i])

# sheetlist = ['Test1']#, 'Test2', 'Test3']

# for w in sheetlist:
#     wb.create_sheet(w)

# for s in sheetlist:

#     values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10)
#     chart = BarChart()
#     chart.add_data(values)

#     ws = wb[s]
#     ws.add_chart(chart, "B5")

# wb.save("SampleChart.xlsx")
